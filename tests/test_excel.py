from dataclasses import replace
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from survey_maker.excel import WorkbookValidationError, create_template, read_answer_rows
from survey_maker.wjx import Choice, Question, QuestionType


QUESTIONS = (
    Question(
        number=1,
        title="地区",
        question_type=QuestionType.SINGLE_CHOICE,
        required=True,
        field_id="div1",
        field_name="q1",
        choices=(Choice("1", "北京"), Choice("2", "上海")),
    ),
    Question(
        number=2,
        title="兴趣",
        question_type=QuestionType.MULTIPLE_CHOICE,
        required=True,
        field_id="div2",
        field_name="q2",
        choices=(Choice("1", "技术"), Choice("2", "设计")),
    ),
    Question(
        number=3,
        title="意见",
        question_type=QuestionType.TEXT,
        required=False,
        field_id="div3",
        field_name="q3",
    ),
)


class ExcelTemplateTests(unittest.TestCase):
    def test_template_uses_question_order_and_reads_valid_rows(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "answers.xlsx"
            create_template(QUESTIONS, path)

            workbook = load_workbook(path)
            answers = workbook["answers"]
            self.assertEqual([answers.cell(1, column).value for column in range(1, 4)], ["Q1", "Q2", "Q3"])
            self.assertIn("question_guide", workbook.sheetnames)
            answers.append(["上海", "技术；2", "请补充服务"])
            workbook.save(path)

            rows = read_answer_rows(path, QUESTIONS)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].excel_row, 2)
            self.assertEqual(rows[0].answers[0].choice_values, ("2",))
            self.assertEqual(rows[0].answers[1].choice_values, ("1", "2"))
            self.assertEqual(rows[0].answers[2].text, "请补充服务")

    def test_required_answer_is_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "answers.xlsx"
            create_template(QUESTIONS, path)
            workbook = load_workbook(path)
            workbook["answers"].append(["", "技术", ""])
            workbook.save(path)

            with self.assertRaisesRegex(WorkbookValidationError, "Q1.*必填"):
                read_answer_rows(path, QUESTIONS)

    def test_reordered_headers_are_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "answers.xlsx"
            create_template(QUESTIONS, path)
            workbook = load_workbook(path)
            workbook["answers"]["A1"] = "Q2"
            workbook.save(path)

            with self.assertRaisesRegex(WorkbookValidationError, "题目顺序"):
                read_answer_rows(path, QUESTIONS)

    def test_too_many_multiple_choice_answers_are_rejected(self) -> None:
        limited_questions = (QUESTIONS[0], replace(QUESTIONS[1], max_choices=1), QUESTIONS[2])
        with TemporaryDirectory() as directory:
            path = Path(directory) / "answers.xlsx"
            create_template(limited_questions, path)
            workbook = load_workbook(path)
            workbook["answers"].append(["北京", "技术；设计", ""])
            workbook.save(path)

            with self.assertRaisesRegex(WorkbookValidationError, "Q2.*最多只能选择 1 项"):
                read_answer_rows(path, limited_questions)

    def test_too_few_multiple_choice_answers_are_rejected(self) -> None:
        limited_questions = (QUESTIONS[0], replace(QUESTIONS[1], min_choices=2), QUESTIONS[2])
        with TemporaryDirectory() as directory:
            path = Path(directory) / "answers.xlsx"
            create_template(limited_questions, path)
            workbook = load_workbook(path)
            workbook["answers"].append(["北京", "技术", ""])
            workbook.save(path)

            with self.assertRaisesRegex(WorkbookValidationError, "Q2.*至少需要选择 2 项"):
                read_answer_rows(path, limited_questions)

    def test_required_unsupported_question_is_only_allowed_for_manual_prepare(self) -> None:
        unsupported = Question(
            number=4,
            title="矩阵题",
            question_type=QuestionType.UNSUPPORTED,
            required=True,
            field_id="div4",
            field_name="q4",
        )
        questions = (*QUESTIONS, unsupported)
        with TemporaryDirectory() as directory:
            path = Path(directory) / "answers.xlsx"
            create_template(questions, path)
            workbook = load_workbook(path)
            workbook["answers"].append(["北京", "技术", "", ""])
            workbook.save(path)

            with self.assertRaisesRegex(WorkbookValidationError, "Q4.*必填且暂不支持"):
                read_answer_rows(path, questions)

            answer_rows = read_answer_rows(path, questions, allow_manual_questions=True)
            self.assertEqual(len(answer_rows), 1)


if __name__ == "__main__":
    unittest.main()