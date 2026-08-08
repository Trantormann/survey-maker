"""Excel 模板生成与逐行答案校验。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .wjx import Choice, Question, QuestionType, iter_choice_labels


_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx"}
_CHOICE_PREFIX = re.compile(r"^[A-Za-z]\s*[.、:：]\s*")
_MULTI_VALUE_SEPARATOR = re.compile(r"[;；]")


class WorkbookValidationError(ValueError):
    """Excel 内容与问卷题目不匹配时抛出。"""


@dataclass(frozen=True)
class AnswerValue:
    question: Question
    text: str | None = None
    choice_values: tuple[str, ...] = ()


@dataclass(frozen=True)
class AnswerRow:
    excel_row: int
    answers: tuple[AnswerValue, ...]


def create_template(questions: Sequence[Question], output_path: str | Path) -> Path:
    """创建可编辑的 Excel 回答模板，并返回生成的路径。"""
    if not questions:
        raise ValueError("未解析到题目，无法创建 Excel 模板。")

    path = _validate_excel_path(output_path)
    workbook = Workbook()
    answers = workbook.active
    answers.title = "answers"
    guide = workbook.create_sheet("question_guide")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, question in enumerate(questions, start=1):
        cell = answers.cell(row=1, column=column_index, value=f"Q{question.number}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.comment = Comment(_question_help(question), "Survey Maker")
        answers.column_dimensions[get_column_letter(column_index)].width = 18

    answers.freeze_panes = "A2"
    answers.auto_filter.ref = f"A1:{get_column_letter(len(questions))}1"

    guide_headers = ("顺序", "答案列", "题目", "类型", "是否必填", "可填写的答案", "填写规则")
    guide.append(guide_headers)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for question in questions:
        choice_text = "\n".join(
            f"{index}. {label}" for index, label in enumerate(iter_choice_labels(question), start=1)
        )
        guide.append(
            (
                question.number,
                f"Q{question.number}",
                question.title,
                _question_type_label(question.question_type),
                "必填" if question.required else "选填",
                choice_text or "-",
                _entry_rule(question),
            )
        )

    guide.freeze_panes = "A2"
    guide.auto_filter.ref = f"A1:G{len(questions) + 1}"
    for row in guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"A": 8, "B": 12, "C": 55, "D": 14, "E": 10, "F": 42, "G": 36}.items():
        guide.column_dimensions[column].width = width

    workbook.save(path)
    return path


def read_answer_rows(excel_path: str | Path, questions: Sequence[Question]) -> list[AnswerRow]:
    """读取 answers 工作表，并将每个非空行验证为一份问卷答案。"""
    if not questions:
        raise ValueError("未解析到题目，无法读取回答。")

    path = _validate_excel_path(excel_path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except OSError as error:
        raise WorkbookValidationError(f"无法读取 Excel 文件：{path}") from error

    try:
        worksheet = workbook["answers"] if "answers" in workbook.sheetnames else workbook.active
        _validate_headers(worksheet, questions)
        rows: list[AnswerRow] = []
        for excel_row, cells in enumerate(
            worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, values_only=True), start=2
        ):
            values = [_cell_text(value) for value in cells]
            if not any(values):
                continue
            if any(values[len(questions) :]):
                raise WorkbookValidationError(f"第 {excel_row} 行存在超出问卷题目数量的非空列。")
            answers = tuple(
                _parse_answer(question, values[index], excel_row)
                for index, question in enumerate(questions)
            )
            rows.append(AnswerRow(excel_row=excel_row, answers=answers))
        return rows
    finally:
        workbook.close()


def _validate_excel_path(path_value: str | Path) -> Path:
    path = Path(path_value)
    if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        raise WorkbookValidationError("仅支持 .xlsx、.xlsm 或 .xltx 格式的 Excel 文件。")
    return path


def _validate_headers(worksheet: object, questions: Sequence[Question]) -> None:
    expected_headers = [f"Q{question.number}" for question in questions]
    actual_headers = [_cell_text(worksheet.cell(row=1, column=index).value) for index in range(1, len(questions) + 1)]
    if actual_headers != expected_headers:
        expected = "、".join(expected_headers)
        actual = "、".join(header or "<空>" for header in actual_headers)
        raise WorkbookValidationError(
            f"answers 工作表第 1 行必须按题目顺序为：{expected}；当前为：{actual}。"
        )


def _parse_answer(question: Question, value: str, excel_row: int) -> AnswerValue:
    if not value:
        if question.required:
            raise WorkbookValidationError(f"第 {excel_row} 行的 Q{question.number} 是必填题，不能留空。")
        return AnswerValue(question=question)

    if question.question_type in {QuestionType.SINGLE_CHOICE, QuestionType.SELECT}:
        return AnswerValue(question=question, choice_values=(_resolve_choice(question, value, excel_row),))

    if question.question_type == QuestionType.MULTIPLE_CHOICE:
        parts = [part.strip() for part in _MULTI_VALUE_SEPARATOR.split(value)]
        if not parts or any(not part for part in parts):
            raise WorkbookValidationError(
                f"第 {excel_row} 行的 Q{question.number} 多选答案请用英文或中文分号分隔，且不能有空项。"
            )
        choices = tuple(_resolve_choice(question, part, excel_row) for part in parts)
        if len(set(choices)) != len(choices):
            raise WorkbookValidationError(f"第 {excel_row} 行的 Q{question.number} 包含重复的多选答案。")
        return AnswerValue(question=question, choice_values=choices)

    if question.question_type == QuestionType.TEXT:
        return AnswerValue(question=question, text=value)

    raise WorkbookValidationError(
        f"第 {excel_row} 行的 Q{question.number} 是暂不支持的题型，必须留空后手动填写。"
    )


def _resolve_choice(question: Question, value: str, excel_row: int) -> str:
    candidate = _normalize(value)
    if candidate.isdigit():
        choice_index = int(candidate)
        if 1 <= choice_index <= len(question.choices):
            return question.choices[choice_index - 1].value

    matches = [choice for choice in question.choices if candidate in _choice_aliases(choice)]
    if len(matches) == 1:
        return matches[0].value
    choices = "；".join(iter_choice_labels(question))
    raise WorkbookValidationError(
        f"第 {excel_row} 行的 Q{question.number} 答案“{value}”不在可选项中。可选项：{choices}"
    )


def _choice_aliases(choice: Choice) -> set[str]:
    label = _normalize(choice.label)
    return {_normalize(choice.value), label, _CHOICE_PREFIX.sub("", label)}


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize(value: str) -> str:
    return " ".join(value.split()).casefold()


def _question_type_label(question_type: QuestionType) -> str:
    return {
        QuestionType.SINGLE_CHOICE: "单选",
        QuestionType.MULTIPLE_CHOICE: "多选",
        QuestionType.SELECT: "下拉选择",
        QuestionType.TEXT: "文本",
        QuestionType.UNSUPPORTED: "暂不支持",
    }[question_type]


def _entry_rule(question: Question) -> str:
    if question.question_type == QuestionType.MULTIPLE_CHOICE:
        return "填写选项全文、选项序号或选项值；多个答案以英文或中文分号分隔。"
    if question.question_type in {QuestionType.SINGLE_CHOICE, QuestionType.SELECT}:
        return "填写选项全文、选项序号或选项值。"
    if question.question_type == QuestionType.TEXT:
        return "填写要输入到文本框中的内容。"
    return "此题型暂不自动填写，请在浏览器中手动完成。"


def _question_help(question: Question) -> str:
    choices = "\n".join(f"{index}. {label}" for index, label in enumerate(iter_choice_labels(question), start=1))
    return "\n".join(
        part
        for part in (
            question.title,
            f"类型：{_question_type_label(question.question_type)}",
            "必填" if question.required else "选填",
            _entry_rule(question),
            choices,
        )
        if part
    )